"""
# 使用示例
excel = CrazyExcel('图片和数据混合.xlsx')
# 批量设置从A1到D5的所有单元格大小为100x100像素
excel.set_cells_size(5, 4, 100, 100)

excel.append(['姓名', '照片', '职业', '收入'])
excel.append(['张三', excel.image('1.jpg', 100, 100), '工程师', 15000])
excel.append(['李四', excel.image('2.jpg', 100, 100), '教师', 12000])
excel.append(['王五', None, '医生', 20000])

excel.save()
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

class CrazyExcel:

    def __init__(self, excel_name: str):
        self.excel_name = excel_name
        self.workbook = openpyxl.Workbook()  # 创建新的工作簿
        self.sheet = self.workbook.active
        self.current_row = 1  # 跟踪当前行号
        # 转换比例配置
        self.px_to_char_width = 7
        self.px_to_point = 1.3333

    def image(self, path: str, width_px: int, height_px: int) -> Image:
        """
        创建一个图片对象
        
        Args:
            path: 图片路径
            width_px: 图片宽度(像素)
            height_px: 图片高度(像素)
        """
        image = Image(path)
        # 直接使用像素值设置图片大小
        image.width = width_px
        image.height = height_px
        return image
    
    def set_cells_size(self, end_row: int, end_col: int, width_px: float, height_px: float):
        """
        批量设置从A1到指定行列的单元格大小
        
        Args:
            end_row: 结束行号
            end_col: 结束列号
            width_px: 列宽(以像素为单位)
            height_px: 行高(以像素为单位)
        """
        # 将像素转换为Excel单位
        width_char = width_px / self.px_to_char_width  # 将像素转换为字符单位
        height_point = height_px / self.px_to_point    # 将像素转换为点单位
        
        # 设置所有指定列的宽度
        for col in range(1, min(end_col + 1, 16384)):  # Excel 的列数限制为 16384
            col_letter = get_column_letter(col)  # 使用 openpyxl 的函数正确处理列名转换
            self.sheet.column_dimensions[col_letter].width = width_char
        
        # 设置所有指定行的高度
        for row in range(1, min(end_row + 1, 1048576)):  # Excel 的行数限制为 1048576
            self.sheet.row_dimensions[row].height = height_point
            
        return self
    
    def append(self, data: list):
        """
        添加一行数据，可以包含字符串、数字和图片对象
        图片对象将被添加到对应的单元格位置
        """
        row_data = []
        image_positions = []  # 保存图片及其位置
        
        for i, cell_content in enumerate(data):
            if isinstance(cell_content, Image):
                # 如果是图片对象，标记位置，稍后添加
                image_positions.append((cell_content, i+1))
                row_data.append(None)  # 在单元格中添加空值
            else:
                # 普通数据直接添加
                row_data.append(cell_content)
        
        # 添加普通数据行
        self.sheet.append(row_data)
        
        # 添加图片
        for img, col_index in image_positions:
            # 计算单元格位置 (例如 'B2', 'C3')
            cell_pos = f'{chr(64 + col_index)}{self.current_row}'
            self.sheet.add_image(img, cell_pos)
        
        self.current_row += 1
        return self
    
    def save(self):
        self.workbook.save(self.excel_name)